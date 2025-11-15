import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# Stock data
sectors_data = {
    "Oil and Gas Exploration Companies": ["MARI", "PPL", "OGDC"],
    "Cement": ["FCCL", "MLCF", "DGKC", "PIOS", "LUCK"],
    "Commercial Banks": ["FABL", "MEBL"],
    "INV. Banks/INV. COS/Securities Cos": ["ENGROH"],
    "Power Generation and Distribution": ["HUBC"],
    "Technology and Communication": ["SYS", "NETSOL", "AVN", "OCTOPUS"],
    "Oil and Gas Marketing Companies": ["SNGP", "PSO"],
    "Fertilizers": ["EFERT"],
    "Automobile Assembler": ["SAZEW", "HCAR", "MTL", "GAL", "GHNI"],
    "Refinery": ["ATRL", "PRL"],
    "Cable and Electrical Goods": ["PAEL"],
    "Pharmaceuticals": ["SEARL"],
    "Food and Personal Care Products": ["FFL", "TOMCL"]
}

# Light colors for each sector
sector_colors = [
    "F0F8FF", "F5F5DC", "E6E6FA", "F0FFF0", "FFF8DC", 
    "F5FFFA", "FDF5E6", "F0F0F0", "FFFACD", "E0FFFF",
    "FFE4E1", "F8F8FF", "FFEFD5"
]

# Company weights from the image
company_weights = {
    "ENGROH": 10.14,
    "LUCK": 9.3,
    "HUBC": 9.29,
    "MEBL": 9.15,
    "MARI": 8.06,
    "OGDC": 7.64,
    "SYS": 6.22,
    "EFERT": 6.19,
    "PPL": 6.18,
    "PSO": 4.88,
    "MLCF": 2.39,
    "DGKC": 2.38,
    "MTL": 2.22,
    "FCCL": 2.22,
    "FABL": 1.9,
    "SAZEW": 1.76,
    "SNGP": 1.59,
    "ATRL": 1.38,
    "PIOS": 1.31,
    "PAEL": 1.29,
    "SEARL": 1.26,
    "GAL": 0.6,
    "GHNI": 0.57,
    "FFL": 0.45,
    "PRL": 0.39,
    "HCAR": 0.39,
    "AVN": 0.31,
    "TOMCL": 0.25,
    "NETSOL": 0.17,
    "OCTOPUS": 0.1
}
company_names = {
    "MARI": "Mari Petroleum Company Limited",
    "PPL": "Pakistan Petroleum Limited",
    "OGDC": "Oil and Gas Development Company Limited",
    "FCCL": "Fauji Cement Company Limited",
    "MLCF": "Maple Leaf Cement Factory Limited",
    "DGKC": "D. G. Khan Cement Company Limited",
    "PIOS": "Pioneer Cement Limited",
    "LUCK": "Lucky Cement Limited",
    "FABL": "Faysal Bank Limited",
    "MEBL": "Muslim Commercial Bank Limited",
    "ENGROH": "Engro Corporation Limited",
    "HUBC": "Hub Power Company Limited",
    "SYS": "Systems Limited",
    "NETSOL": "NetSol Technologies Limited",
    "AVN": "Avanceon Limited",
    "OCTOPUS": "Octopus Digital Limited",
    "SNGP": "Sui Northern Gas Pipelines Limited",
    "PSO": "Pakistan State Oil Company Limited",
    "EFERT": "Engro Fertilizers Limited",
    "SAZEW": "Sazgar Engineering Works Limited",
    "HCAR": "Honda Cars Pakistan Limited",
    "MTL": "Millat Tractors Limited",
    "GAL": "Ghandhara Automobile Limited",
    "GHNI": "Ghandhara Nissan Limited",
    "ATRL": "Attock Refinery Limited",
    "PRL": "Pakistan Refinery Limited",
    "PAEL": "Pakistan Cables Limited",
    "SEARL": "The Searle Company Limited",
    "FFL": "Fauji Foods Limited",
    "TOMCL": "Tomato Pakistan Limited"
}

# Create workbook
wb = openpyxl.Workbook()

# Header style
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

top_border = Border(top=Side(style='thin'))
bottom_border = Border(bottom=Side(style='thin'))
left_border = Border(left=Side(style='thin'))
right_border = Border(right=Side(style='thin'))
top_left_border = Border(top=Side(style='thin'), left=Side(style='thin'))
top_right_border = Border(top=Side(style='thin'), right=Side(style='thin'))
bottom_left_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
bottom_right_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))

# Worksheet 1: Sectors Summary
ws1 = wb.active
ws1.title = "Sectors Summary"

ws1['A1'] = "Sector"
ws1['B1'] = "Total Companies"
ws1['A1'].font = header_font
ws1['A1'].fill = header_fill
ws1['A1'].border = thin_border
ws1['B1'].font = header_font
ws1['B1'].fill = header_fill
ws1['B1'].border = thin_border

row = 2
color_index = 0
for sector, companies in sectors_data.items():
    sector_fill = PatternFill(start_color=sector_colors[color_index], end_color=sector_colors[color_index], fill_type="solid")
    ws1[f'A{row}'] = sector
    ws1[f'B{row}'] = len(companies)
    ws1[f'A{row}'].fill = sector_fill
    ws1[f'B{row}'].fill = sector_fill
    
    # Add borders around the sector row
    ws1[f'A{row}'].border = Border(top=Side(style='thin'), left=Side(style='thin'), bottom=Side(style='thin'))
    ws1[f'B{row}'].border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
    
    row += 1
    color_index += 1

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 20

# Worksheet 2: Company Details
ws2 = wb.create_sheet(title="Company Details")

# Headers outside the table (row 1)
ws2['A1'] = "Sector"
ws2['B1'] = "Company Short Name"
ws2['C1'] = "Company Full Name"
ws2['D1'] = "Weight"

# Format headers
ws2['A1'].font = header_font
ws2['A1'].fill = header_fill
ws2['B1'].font = header_font
ws2['B1'].fill = header_fill
ws2['C1'].font = header_font
ws2['C1'].fill = header_fill
ws2['D1'].font = header_font
ws2['D1'].fill = header_fill

# Data starts from row 3 (leaving row 2 empty for separation)
row = 3
color_index = 0
for sector, companies in sectors_data.items():
    sector_fill = PatternFill(start_color=sector_colors[color_index], end_color=sector_colors[color_index], fill_type="solid")
    for company in companies:
        ws2[f'A{row}'] = sector
        ws2[f'B{row}'] = company
        ws2[f'C{row}'] = company_names.get(company, f"{company} Limited")
        ws2[f'D{row}'] = company_weights.get(company, 0)
        
        # Apply sector colors
        ws2[f'A{row}'].fill = sector_fill
        ws2[f'B{row}'].fill = sector_fill
        ws2[f'C{row}'].fill = sector_fill
        ws2[f'D{row}'].fill = sector_fill
        
        row += 1
    color_index += 1

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 40
ws2.column_dimensions['D'].width = 15

# Apply autofilter only to the data table (row 3 onwards)
ws2.auto_filter.ref = f"A3:D{row-1}"

# Save file
wb.save('/Users/muneebahmad/personal/stocks/stocks_data.xlsx')
print("Excel file created with two separate worksheets at /Users/muneebahmad/personal/stocks/stocks_data.xlsx")
